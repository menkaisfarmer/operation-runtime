from typing import Any, Dict, List, Optional
from uuid import uuid4

from core.target import Target
from core.action import Action, Delete
from .base import BaseAdapter


class ExcelAdapter(BaseAdapter):
    """Excel ファイル用 Adapter（openpyxl使用）"""

    def __init__(self, file_path: str, sheet_name: str = "Sheet1"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
        self._transactions = {}

    def connect(self) -> None:
        """Excel ファイルを開く"""
        try:
            import openpyxl

            self.wb = openpyxl.load_workbook(self.file_path)
            self.ws = self.wb[self.sheet_name]
        except FileNotFoundError:
            import openpyxl

            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = self.sheet_name

    def disconnect(self) -> None:
        """Excel ファイルを閉じる"""
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws = None

    def read(
        self, target: Optional[Target] = None
    ) -> List[Dict[str, Any]]:
        """Excel からデータを読み取る"""
        if not self.is_connected():
            raise RuntimeError("Adapter not connected")

        records = []
        rows = self.ws.iter_rows(values_only=False)

        # ヘッダー行を取得
        header_row = next(rows, None)
        if header_row is None:
            return records

        headers = [cell.value for cell in header_row]

        # データ行を読み取る
        for row in rows:
            record = {}
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    record[headers[idx]] = cell.value
            if any(record.values()):  # 空行を除外
                records.append(record)

        if target is not None:
            records = target.filter(records)

        return records

    def write(self, records: List[Dict[str, Any]]) -> bool:
        """Excel にデータを書き込む"""
        if not self.is_connected():
            raise RuntimeError("Adapter not connected")

        try:
            # シートをクリア
            for row in self.ws.iter_rows():
                for cell in row:
                    cell.value = None

            if not records:
                return True

            # ヘッダー行を書き込む
            headers = list(records[0].keys())
            for col_idx, header in enumerate(headers, 1):
                self.ws.cell(row=1, column=col_idx, value=header)

            # データ行を書き込む
            for row_idx, record in enumerate(records, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = record.get(header)
                    self.ws.cell(row=row_idx, column=col_idx, value=value)

            self.wb.save(self.file_path)
            return True
        except Exception as e:
            raise RuntimeError(f"Write failed: {str(e)}")

    def apply_action(
        self, record: Dict[str, Any], action: Action
    ) -> Dict[str, Any]:
        """アクションを適用"""
        if isinstance(action, Delete):
            return None
        return action.apply(record)

    def begin_transaction(self) -> str:
        """トランザクション開始"""
        if not self.is_connected():
            raise RuntimeError("Adapter not connected")

        tx_id = str(uuid4())
        # スナップショット用にデータを読み取る
        snapshot = self.read()
        self._transactions[tx_id] = snapshot
        return tx_id

    def commit_transaction(self, tx_id: str) -> bool:
        """トランザクションコミット"""
        if tx_id not in self._transactions:
            return False

        del self._transactions[tx_id]
        return True

    def rollback_transaction(self, tx_id: str) -> bool:
        """トランザクションロールバック"""
        if tx_id not in self._transactions:
            return False

        try:
            snapshot = self._transactions[tx_id]
            self.write(snapshot)
            del self._transactions[tx_id]
            return True
        except Exception as e:
            raise RuntimeError(f"Rollback failed: {str(e)}")

    def is_connected(self) -> bool:
        """接続状態を確認"""
        return self.wb is not None
